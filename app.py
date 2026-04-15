"""
Bluepeak Energy  |  Wattix Tools
Two-tab Streamlit app:
  Tab 1 — Excel Generator  (Wattix CSV/XLSX → formatted Excel report)
  Tab 2 — PPT Generator    (Bluepeak Excel output → PowerPoint presentation)
"""

import os, re, copy, zipfile, warnings, tempfile
warnings.filterwarnings("ignore")

from datetime import datetime
from io import BytesIO
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from lxml import etree
import streamlit as st

SCRIPT_DIR   = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_PATH = os.path.join(SCRIPT_DIR, "Bluepeak_Template.pptx")


# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL GENERATOR  —  core logic (no tkinter)
# ══════════════════════════════════════════════════════════════════════════════

C_DARK="1B4332"; C_MID="40916C"; C_LIGHT="D8F3DC"; C_TOTAL="B7E4C7"
C_GRAY="F8F9FA"; C_WHITE="FFFFFF"; C_BORDER="AAAAAA"; C_BLUE="0000FF"
FMT_INT='#,##0'; FMT_PCT='0.0%'
MONTHS=["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]

def col(n): return get_column_letter(n)
def fill(c=C_DARK): return PatternFill("solid",fgColor=c)
def border():
    s=Side(style="thin",color=C_BORDER); return Border(left=s,right=s,top=s,bottom=s)
def aln(h="left",v="center"): return Alignment(horizontal=h,vertical=v)
def fnt(bold=False,size=10,color="000000"): return Font(name="Arial",bold=bold,size=size,color=color)
def hdr(ws,r,c1,c2,text,bg=C_DARK,fg=C_WHITE,size=11):
    ws.merge_cells(start_row=r,start_column=c1,end_row=r,end_column=c2)
    cell=ws.cell(row=r,column=c1,value=text)
    cell.fill=fill(bg); cell.font=Font(name="Arial",bold=True,color=fg,size=size); cell.alignment=aln()

def sname(name, wb=None):
    name = re.sub(r'[\\/*?:\[\]]', '', name)
    existing = {ws.title for ws in wb.worksheets} if wb is not None else set()
    m = re.search(r'\s[–\-]\s\S.*$', name)
    if m and len(name) > 31:
        body = name[:m.start()]; suffix = name[m.start():]
        base = (body[:31 - len(suffix)] + suffix)[:31]
    else:
        body = None; suffix = None; base = name[:31]
    if base not in existing: return base
    for i in range(1, 1000):
        n_str = str(i)
        if body is not None and suffix is not None:
            candidate = (body[:31 - len(suffix) - len(n_str)] + n_str + suffix)[:31]
        else:
            candidate = base[:31 - len(n_str)] + n_str
        if candidate not in existing: return candidate
    raise ValueError(f"Cannot create unique sheet name for: {name!r}")

def widths(ws, wl):
    for i, w in enumerate(wl, 1): ws.column_dimensions[col(i)].width = w

def read_file(path, **kw):
    return pd.read_csv(path, **kw) if path.lower().endswith('.csv') else pd.read_excel(path, **kw)

def peek_columns(path):
    try: return [c.lower().strip() for c in read_file(path, nrows=0).columns]
    except: return []

def is_monthly(path):
    cols = peek_columns(path); return 'year' in cols and 'month' in cols

def scenario_key(filename):
    name = re.sub(r'\.(csv|xlsx)$', '', filename, flags=re.I)
    m = re.search(r'[Ss]olution\s*(\d+)', name)
    if m: return int(m.group(1))
    parts = [p for p in name.split('_') if p]
    return parts[-1] if parts else name

def _norm_key(s):
    s = re.sub(r'\.(csv|xlsx)$', '', str(s), flags=re.I).lower()
    for word in ['wattix','stored','bess','opgeslagen','load','production','gen']:
        s = re.sub(r'\b'+word+r'\b', '', s)
    return re.sub(r'[^a-z0-9]', '', s)

def _resolve_stored(stf, all_keys):
    result = {}; claimed = set(); unmatched = []
    for k in all_keys:
        nk = _norm_key(str(k))
        if nk in stf: result[k] = stf[nk]; claimed.add(nk)
        else: unmatched.append(k)
    unclaimed = {sk: sv for sk, sv in stf.items() if sk not in claimed}
    for k in unmatched:
        nk = _norm_key(str(k))
        candidates = {sk: sv for sk, sv in unclaimed.items() if sk.startswith(nk)}
        if len(candidates) == 1:
            sk = next(iter(candidates)); result[k] = candidates[sk]; claimed.add(sk)
    return result

def classify_files(paths):
    lf, pf, gf, stf, sf = {}, {}, {}, {}, {}
    for path in paths:
        f = os.path.basename(path); fl = f.lower()
        if 'template' in fl: continue
        if not (fl.endswith('.csv') or fl.endswith('.xlsx')): continue
        if ('stored' in fl or 'bess' in fl or 'opgeslagen' in fl) and 'load' not in fl and 'production' not in fl:
            stf[_norm_key(f)] = path; continue
        if 'wattix' not in fl: continue
        key = scenario_key(f)
        if 'gen' in fl and 'load' not in fl and 'production' not in fl: gf[key] = path
        elif 'production' in fl:
            if is_monthly(path): pf[key] = path
        elif 'load' in fl:
            if is_monthly(path): lf[key] = path
            else:
                rem = re.sub(r'(?i)wattix[^_]*_', '', re.sub(r'\.(csv|xlsx)$', '', f, flags=re.I))
                sn = rem.split('_')[0].capitalize() or f; sf[sn] = path
    return lf, pf, gf, stf, sf

ECOLS_L = ['year','month','load_kwh','source_grid_kwh','source_production_kwh','gen_kwh','source_gen_kwh','source_grid_only_kwh']
ECOLS_P = ['year','month','production_kwh','self_consumption_direct_kwh','stored_kwh','curtailed_kwh','potentiele_opwek']

def _best_year(df):
    if df.empty or 'year' not in df.columns or 'month' not in df.columns: return None
    for y in sorted(df['year'].unique(), reverse=True):
        if df[df['year']==y]['month'].nunique() == 12: return int(y)
    return None

def _filter_year(df, year):
    if year is None or df.empty: return df
    return df[df['year']==year].copy()

def load_solution(key, load_path, prod_path, gen_path, stored_path, site_path):
    sol = {'num': key, 'missing': []}
    if load_path:
        ld = read_file(load_path); ld.columns = [c.lower().strip() for c in ld.columns]
    else:
        sol['missing'].append('load file'); ld = pd.DataFrame(columns=ECOLS_L)
    if prod_path:
        pd_ = read_file(prod_path); pd_.columns = [c.lower().strip() for c in pd_.columns]
    else:
        sol['missing'].append('production file'); pd_ = pd.DataFrame(columns=ECOLS_P)
    if gen_path:
        gd = read_file(gen_path); gd.columns = [c.lower().strip() for c in gd.columns]
        gm = gd.groupby(['year','month'])['value'].sum().reset_index(); gm.columns = ['year','month','gen_kwh']
    else:
        gd = pd.DataFrame(columns=['year','month','day','date','value']); gm = pd.DataFrame(columns=['year','month','gen_kwh'])
    if stored_path:
        sd = read_file(stored_path); sd.columns = [c.lower().strip() for c in sd.columns]
        sm = sd.groupby(['year','month'])['value'].sum().reset_index(); sm.columns = ['year','month','bess_kwh']
    else:
        sm = pd.DataFrame(columns=['year','month','bess_kwh'])
    if not ld.empty:
        ld = ld.merge(gm, on=['year','month'], how='left')
        ld['gen_kwh'] = ld['gen_kwh'].fillna(0).infer_objects(copy=False)
        ld['source_gen_kwh'] = ld['gen_kwh']
        if not sm.empty:
            ld = ld.merge(sm, on=['year','month'], how='left')
            ld['bess_kwh'] = ld['bess_kwh'].fillna(0).infer_objects(copy=False)
            ld['source_battery_kwh'] = ld['bess_kwh']
            ld['source_grid_only_kwh'] = (ld['source_grid_kwh']-ld['gen_kwh']-ld['bess_kwh']).clip(lower=0)
        elif not pd_.empty and 'stored_kwh' in pd_.columns:
            prod_stored = pd_[['year','month','stored_kwh']].groupby(['year','month'])['stored_kwh'].sum().reset_index()
            ld = ld.merge(prod_stored, on=['year','month'], how='left')
            ld['stored_kwh'] = ld['stored_kwh'].fillna(0).infer_objects(copy=False)
            ld['source_battery_kwh'] = ld['stored_kwh']
            ld['source_production_kwh'] = (ld['source_production_kwh']-ld['stored_kwh']).clip(lower=0)
            ld['source_grid_only_kwh'] = (ld['source_grid_kwh']-ld['gen_kwh']).clip(lower=0)
        else:
            ld['source_battery_kwh'] = 0
            ld['source_grid_only_kwh'] = (ld['source_grid_kwh']-ld['gen_kwh']).clip(lower=0)
    else:
        for c in ['gen_kwh','source_gen_kwh','source_grid_only_kwh','source_battery_kwh']: ld[c] = 0
    if not pd_.empty and 'production_kwh' in pd_.columns:
        pd_['potentiele_opwek'] = pd_['production_kwh'] + pd_['curtailed_kwh']
        if not sm.empty:
            pd_ = pd_.merge(sm, on=['year','month'], how='left')
            pd_['bess_kwh'] = pd_['bess_kwh'].fillna(0).infer_objects(copy=False)
        else:
            pd_['bess_kwh'] = pd_['stored_kwh'].fillna(0)
    else:
        pd_['potentiele_opwek'] = 0; pd_['bess_kwh'] = 0
    best = _best_year(ld) or _best_year(pd_)
    if best:
        ld = _filter_year(ld, best); pd_ = _filter_year(pd_, best)
        gd = _filter_year(gd, best) if not gd.empty else gd
    site_df = None; site_name = ''
    if site_path:
        is_csv = site_path.lower().endswith('.csv')
        if is_csv:
            df = read_file(site_path); df.columns = [c.lower().strip() for c in df.columns]
            tc = next((c for c in df.columns if any(k in c for k in ('tijd','time','stamp'))), df.columns[0])
            lc = next((c for c in df.columns if 'load' in c or 'kwh' in c), df.columns[1])
            df = df.rename(columns={tc:'timestamp', lc:'load_kwh'})[['timestamp','load_kwh']].copy()
            site_df = df
        else:
            df = pd.read_excel(site_path); df = df.iloc[1:].reset_index(drop=True)
            df.columns = ['timestamp','load_kwh'] + list(range(df.shape[1]-2))
            site_df = df[['timestamp','load_kwh']].copy()
        site_df = site_df.dropna(subset=['timestamp'])
        site_df['load_kwh'] = pd.to_numeric(site_df['load_kwh'], errors='coerce')
        site_name = os.path.basename(site_path)
        site_name = re.sub(r'\.(csv|xlsx)$', '', site_name, flags=re.I)
        site_name = re.sub(r'(?i)wattix[^_]*_', '', site_name).split('_')[0].capitalize()
        sol['site_summary'] = {'total_kwh': site_df['load_kwh'].sum()}; sol['site_name'] = site_name
    sol.update({'load_df': ld, 'prod_df': pd_, 'gen_df': gd, 'site_df': site_df}); return sol

def ann(sol, c, dk='load_df'):
    df = sol.get(dk, pd.DataFrame()); return (df[c].sum()/1000) if (not df.empty and c in df.columns) else 0
def pct(a, b): return a/b if b else 0

def build_excel(solutions, project, location, date, outpath):
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    _summary(wb, solutions, project, location, date)
    _config(wb, solutions, project, location, date)
    for sol in solutions:
        _load_tab(wb, sol); _production_tab(wb, sol)
    wb.save(outpath)

def _summary(wb, solutions, project, location, date):
    ws = wb.create_sheet("📋 Summary"); ws.sheet_view.showGridLines = False; ns = len(solutions)
    hdr(ws,1,1,2+ns,f"  BLUEPEAK ENERGY  |  Wattix Scenario Analysis",size=13)
    hdr(ws,2,1,2+ns,f"  {project}  |  {location}  |  {date}",C_MID,size=10)
    ws.row_dimensions[1].height=28; ws.row_dimensions[2].height=18; ws.row_dimensions[3].height=6
    r=4
    for sol in solutions:
        if sol.get('missing'):
            hdr(ws,r,1,2+ns,f"  ⚠  Scenario '{sol['num']}': missing {', '.join(sol['missing'])}","CC3300","FFFFFF",9)
            ws.row_dimensions[r].height=14; r+=1
    for ci,h in enumerate(["Metric","Unit"]+[str(s['num']) for s in solutions],1):
        c=ws.cell(row=r,column=ci,value=h); c.fill=fill(); c.border=border()
        c.font=Font(name="Arial",bold=True,color=C_WHITE,size=10)
        c.alignment=aln("center") if ci>2 else aln()
    ws.row_dimensions[r].height=20; r+=1
    sections=[
        ("⚡ Load",[("Annual consumption","MWh",[ann(s,'load_kwh') for s in solutions],FMT_INT),
            ("From grid","MWh",[ann(s,'source_grid_only_kwh') for s in solutions],FMT_INT),
            ("From genset","MWh",[ann(s,'source_gen_kwh') for s in solutions],FMT_INT),
            ("From battery","MWh",[ann(s,'source_battery_kwh') for s in solutions],FMT_INT),
            ("From solar","MWh",[ann(s,'source_production_kwh') for s in solutions],FMT_INT)]),
        ("☀️ Production",[("Total production","MWh",[ann(s,'production_kwh','prod_df') for s in solutions],FMT_INT),
            ("Self-consumed","MWh",[ann(s,'self_consumption_direct_kwh','prod_df') for s in solutions],FMT_INT),
            ("Stored in BESS","MWh",[ann(s,'stored_kwh','prod_df') for s in solutions],FMT_INT),
            ("Curtailed","MWh",[ann(s,'curtailed_kwh','prod_df') for s in solutions],FMT_INT),
            ("Potential production","MWh",[ann(s,'potentiele_opwek','prod_df') for s in solutions],FMT_INT)]),
        ("📊 Ratios",[("Grid dependency","%",[pct(ann(s,'source_grid_only_kwh'),ann(s,'load_kwh')) for s in solutions],FMT_PCT),
            ("Solar self-sufficiency","%",[pct(ann(s,'source_production_kwh'),ann(s,'load_kwh')) for s in solutions],FMT_PCT),
            ("Self-consumption ratio","%",[pct(ann(s,'self_consumption_direct_kwh','prod_df'),ann(s,'production_kwh','prod_df')) for s in solutions],FMT_PCT),
            ("Curtailment rate","%",[pct(ann(s,'curtailed_kwh','prod_df'),ann(s,'potentiele_opwek','prod_df')) for s in solutions],FMT_PCT)])]
    alt=False
    for sec_label,rows in sections:
        hdr(ws,r,1,2+ns,f"  {sec_label}",C_MID); ws.row_dimensions[r].height=18; r+=1
        for label,unit,vals,fmt in rows:
            bg=C_LIGHT if alt else C_WHITE
            for ci,v in enumerate([label,unit]+vals,1):
                c=ws.cell(row=r,column=ci,value=v if ci<=2 else round(v,6))
                c.fill=fill(bg); c.border=border(); c.font=fnt(color="666666") if ci==2 else fnt()
                c.number_format=fmt if ci>2 else "@"; c.alignment=aln("right") if ci>2 else (aln("center") if ci==2 else aln())
            ws.row_dimensions[r].height=16; alt=not alt; r+=1
    CFG="'⚙️ Config'"; FMT_EUR='#,##0 "€"'; FMT_YR='0.0 "yr"'
    def cfg_cell(i,offset): return f"{CFG}!B{13+i*16+offset}"
    solar_mwh=[ann(s,'source_production_kwh')+ann(s,'stored_kwh','prod_df') for s in solutions]
    fin_rows=[
        ("PPA / electricity price","€/MWh",[f"={cfg_cell(i,10)}" for i in range(ns)],FMT_EUR),
        ("Total CapEx (PV + BESS)","€",[f"={cfg_cell(i,8)}+{cfg_cell(i,9)}" for i in range(ns)],FMT_EUR),
        ("Annualised PV cost","€/yr",[f"=IFERROR({cfg_cell(i,8)}/IF({cfg_cell(i,12)}>0,{cfg_cell(i,12)},1),\"\")" for i in range(ns)],FMT_EUR),
        ("Annualised BESS cost","€/yr",[f"=IFERROR({cfg_cell(i,9)}/IF({cfg_cell(i,13)}>0,{cfg_cell(i,13)},1),\"\")" for i in range(ns)],FMT_EUR),
        ("Annual savings (solar × PPA)","€/yr",[f"={cfg_cell(i,10)}*{round(solar_mwh[i],4)}" for i in range(ns)],FMT_EUR),
        ("Simple payback (CapEx / savings)","yr",[f"=IFERROR(({cfg_cell(i,8)}+{cfg_cell(i,9)})/({cfg_cell(i,10)}*{round(solar_mwh[i],4)}),\"\")" for i in range(ns)],FMT_YR),
    ]
    hdr(ws,r,1,2+ns,"  💶 Financials  (auto-calculates once Config values are filled in)",C_MID)
    ws.row_dimensions[r].height=18; r+=1
    for label,unit,vals,fmt in fin_rows:
        bg=C_LIGHT if alt else C_WHITE
        c=ws.cell(row=r,column=1,value=label); c.fill=fill(bg); c.border=border(); c.font=fnt()
        c=ws.cell(row=r,column=2,value=unit); c.fill=fill(bg); c.border=border(); c.font=fnt(color="666666"); c.alignment=aln("center")
        for i,formula in enumerate(vals):
            c=ws.cell(row=r,column=3+i,value=formula)
            c.fill=fill(bg); c.border=border(); c.font=fnt(); c.number_format=fmt; c.alignment=aln("right")
        ws.row_dimensions[r].height=16; alt=not alt; r+=1
    widths(ws,[36,10]+[18]*ns); ws.freeze_panes="A5"

def _config(wb, solutions, project, location, date):
    ws=wb.create_sheet("⚙️ Config"); ws.sheet_view.showGridLines=False
    hdr(ws,1,1,4,"  PROJECT CONFIGURATION  |  Fill in blue cells",size=12)
    hdr(ws,2,1,4,"  These inputs feed the Summary and the PowerPoint",C_MID,size=10)
    ws.row_dimensions[1].height=26; ws.row_dimensions[2].height=16; ws.row_dimensions[3].height=6
    def irow(r,label,val='',note=''):
        ws.cell(row=r,column=1,value=label).font=fnt(bold=True)
        c=ws.cell(row=r,column=2,value=val); c.font=Font(name="Arial",color=C_BLUE,size=10)
        if note: ws.cell(row=r,column=3,value=note).font=Font(name="Arial",color="888888",size=9,italic=True)
        for ci in range(1,5): ws.cell(row=r,column=ci).border=border(); ws.cell(row=r,column=ci).fill=fill(C_GRAY if ci==2 else C_WHITE)
        ws.row_dimensions[r].height=15
    hdr(ws,4,1,4,"  General",C_MID)
    for r,(l,v) in enumerate([("Project name",project),("Location",location),("Report date",date),("Prepared by",""),("Client name",""),("Client contact","")],5): irow(r,l,v)
    r=12
    for sol in solutions:
        hdr(ws,r,1,4,f"  Scenario: {sol['num']}",C_MID); r+=1
        for l,n in [("Scenario label","Short display name"),("PV capacity (kWp)","DC peak"),("PV system (kVa)","AC inverter"),
            ("BESS capacity (kW)","Power rating"),("BESS capacity (kWh)","Energy rating"),("Generator (kVA)","Backup generator"),
            ("Grid connection (kVA)","Existing connection"),("Number of panels",""),("PV CapEx (€)","Total installed cost"),
            ("BESS CapEx (€)","Total installed cost"),("PPA price (€/MWh)","Electricity price"),("EPC (%)",""),
            ("PV lifetime (years)",""),("BESS lifetime (years)","")]: irow(r,l,note=n); r+=1
        r+=1
    widths(ws,[28,22,36,8])

def _load_tab(wb, sol):
    n=sol['num']; ws=wb.create_sheet(sname(f"Sc {n} – Load",wb)); ws.sheet_view.showGridLines=False
    hdr(ws,1,1,10,f"  SCENARIO {n}  |  Monthly Load (kWh)",size=12)
    ws.row_dimensions[1].height=26; ws.row_dimensions[2].height=6
    heads=["Month","Total Load","From Grid","From Genset","From Battery","From Solar","Grid %","Genset %","Battery %","Solar %"]
    cw=[15,14,13,13,13,13,10,10,10,10]
    for ci,(h,w) in enumerate(zip(heads,cw),1):
        c=ws.cell(row=3,column=ci,value=h); c.fill=fill(); c.border=border()
        c.font=Font(name="Arial",bold=True,color=C_WHITE,size=10); c.alignment=aln("center")
    ws.row_dimensions[3].height=28
    ld=sol['load_df']
    if not ld.empty:
        agg=ld.groupby('month').agg({'load_kwh':'sum','source_grid_only_kwh':'sum','source_gen_kwh':'sum','source_battery_kwh':'sum','source_production_kwh':'sum'}).reindex(range(1,13),fill_value=0).reset_index()
    else:
        agg=pd.DataFrame({'month':range(1,13),'load_kwh':0,'source_grid_only_kwh':0,'source_gen_kwh':0,'source_battery_kwh':0,'source_production_kwh':0})
    r=4; ds=r; alt=False
    for _,rd in agg.iterrows():
        m=int(rd['month']); bg=C_LIGHT if alt else C_WHITE
        lo=rd['load_kwh'] or 0; gr=rd['source_grid_only_kwh'] or 0; ge=rd['source_gen_kwh'] or 0; ba=rd['source_battery_kwh'] or 0; so=rd['source_production_kwh'] or 0
        vals=[MONTHS[m-1] if 1<=m<=12 else str(m),lo,gr,ge,ba,so,gr/lo if lo else 0,ge/lo if lo else 0,ba/lo if lo else 0,so/lo if lo else 0]
        fmts=[None]+[FMT_INT]*5+[FMT_PCT]*4
        for ci,(v,fmt) in enumerate(zip(vals,fmts),1):
            c=ws.cell(row=r,column=ci,value=v); c.fill=fill(bg); c.font=fnt(); c.border=border()
            c.alignment=aln("right") if fmt else aln()
            if fmt: c.number_format=fmt
        ws.row_dimensions[r].height=15; alt=not alt; r+=1
    de=r-1; ws.row_dimensions[r].height=18
    ws.cell(row=r,column=1,value="TOTAL (kWh)").font=fnt(bold=True); ws.cell(row=r,column=1).fill=fill(C_TOTAL); ws.cell(row=r,column=1).border=border()
    for ci in range(2,7):
        c=ws.cell(row=r,column=ci); c.value=f"=SUM({col(ci)}{ds}:{col(ci)}{de})"
        c.number_format=FMT_INT; c.font=fnt(bold=True); c.fill=fill(C_TOTAL); c.border=border(); c.alignment=aln("right")
    for ci,sc in [(7,3),(8,4),(9,5),(10,6)]:
        c=ws.cell(row=r,column=ci); c.value=f"={col(sc)}{r}/{col(2)}{r}"
        c.number_format=FMT_PCT; c.font=fnt(bold=True); c.fill=fill(C_TOTAL); c.border=border(); c.alignment=aln("right")
    tr=r; r+=1; ws.row_dimensions[r].height=15
    ws.cell(row=r,column=1,value="TOTAL (MWh)").font=fnt(bold=True,color="444444"); ws.cell(row=r,column=1).fill=fill(C_GRAY); ws.cell(row=r,column=1).border=border()
    for ci in range(2,7):
        c=ws.cell(row=r,column=ci); c.value=f"={col(ci)}{tr}/1000"; c.number_format='#,##0.0'
        c.font=fnt(bold=True,color="444444"); c.fill=fill(C_GRAY); c.border=border(); c.alignment=aln("right")
    for ci in [7,8,9,10]: ws.cell(row=r,column=ci).fill=fill(C_GRAY); ws.cell(row=r,column=ci).border=border()
    chart_anchor_row=r+2
    c1=BarChart(); c1.type="col"; c1.grouping="stacked"; c1.overlap=100
    c1.title=f"Scenario {n} – Monthly Load Mix"; c1.y_axis.title="kWh"; c1.y_axis.numFmt='#,##0'
    c1.x_axis.title="Month"; c1.style=10; c1.width=17; c1.height=12; c1.legend.position='b'
    d1=Reference(ws,min_col=3,max_col=6,min_row=3,max_row=de); c1.add_data(d1,titles_from_data=True)
    cats=Reference(ws,min_col=1,min_row=ds,max_row=de); c1.set_categories(cats)
    for ser,color in zip(c1.series,["4472C4","ED7D31","A9C4E4","40916C"]):
        ser.graphicalProperties.solidFill=color; ser.graphicalProperties.line.solidFill=color
    ws.add_chart(c1,f"A{chart_anchor_row}"); widths(ws,cw); ws.freeze_panes="B4"

def _production_tab(wb, sol):
    n=sol['num']; ws=wb.create_sheet(sname(f"Sc {n} – Production",wb)); ws.sheet_view.showGridLines=False
    hdr(ws,1,1,6,f"  SCENARIO {n}  |  Monthly Production (kWh)",size=12)
    ws.row_dimensions[1].height=26; ws.row_dimensions[2].height=6
    heads=["Month","Production","Self Consumption","Stored BESS","Curtailed","Potential"]; cw=[15,14,16,13,11,14]
    for ci,(h,w) in enumerate(zip(heads,cw),1):
        c=ws.cell(row=3,column=ci,value=h); c.fill=fill(); c.border=border()
        c.font=Font(name="Arial",bold=True,color=C_WHITE,size=10); c.alignment=aln("center")
    ws.row_dimensions[3].height=28
    pd_=sol['prod_df']
    if not pd_.empty:
        agg=pd_.groupby('month').agg({'production_kwh':'sum','self_consumption_direct_kwh':'sum','bess_kwh':'sum','curtailed_kwh':'sum','potentiele_opwek':'sum'}).reindex(range(1,13),fill_value=0).reset_index()
    else:
        agg=pd.DataFrame({'month':range(1,13),'production_kwh':0,'self_consumption_direct_kwh':0,'bess_kwh':0,'curtailed_kwh':0,'potentiele_opwek':0})
    r=4; ds=r; alt=False
    for _,rd in agg.iterrows():
        m=int(rd['month']); bg=C_LIGHT if alt else C_WHITE
        pr=rd['production_kwh'] or 0; sc_=rd['self_consumption_direct_kwh'] or 0; st=rd['bess_kwh'] or 0; cu=rd['curtailed_kwh'] or 0; po=rd['potentiele_opwek'] or 0
        vals=[MONTHS[m-1] if 1<=m<=12 else str(m),pr,sc_,st,cu,po]; fmts=[None]+[FMT_INT]*5
        for ci,(v,fmt) in enumerate(zip(vals,fmts),1):
            c=ws.cell(row=r,column=ci,value=v); c.fill=fill(bg); c.font=fnt(); c.border=border()
            c.alignment=aln("right") if fmt else aln()
            if fmt: c.number_format=fmt
        ws.row_dimensions[r].height=15; alt=not alt; r+=1
    de=r-1; ws.row_dimensions[r].height=18
    ws.cell(row=r,column=1,value="TOTAL (kWh)").font=fnt(bold=True); ws.cell(row=r,column=1).fill=fill(C_TOTAL); ws.cell(row=r,column=1).border=border()
    for ci in range(2,7):
        c=ws.cell(row=r,column=ci); c.value=f"=SUM({col(ci)}{ds}:{col(ci)}{de})"
        c.number_format=FMT_INT; c.font=fnt(bold=True); c.fill=fill(C_TOTAL); c.border=border(); c.alignment=aln("right")
    tr=r; r+=1; ws.row_dimensions[r].height=15
    ws.cell(row=r,column=1,value="TOTAL (MWh)").font=fnt(bold=True,color="444444"); ws.cell(row=r,column=1).fill=fill(C_GRAY); ws.cell(row=r,column=1).border=border()
    for ci in range(2,7):
        c=ws.cell(row=r,column=ci); c.value=f"={col(ci)}{tr}/1000"; c.number_format='#,##0.0'
        c.font=fnt(bold=True,color="444444"); c.fill=fill(C_GRAY); c.border=border(); c.alignment=aln("right")
    chart_anchor_row=r+2
    c2=BarChart(); c2.type="col"; c2.grouping="stacked"; c2.overlap=100
    c2.title=f"Scenario {n} – Monthly Production"; c2.y_axis.title="kWh"; c2.y_axis.numFmt='#,##0'
    c2.x_axis.title="Month"; c2.style=10; c2.width=17; c2.height=12; c2.legend.position='b'
    d2=Reference(ws,min_col=3,max_col=5,min_row=3,max_row=de); c2.add_data(d2,titles_from_data=True)
    cats=Reference(ws,min_col=1,min_row=ds,max_row=de); c2.set_categories(cats)
    for ser,color in zip(c2.series,["1B4332","40916C","95D5B2"]):
        ser.graphicalProperties.solidFill=color; ser.graphicalProperties.line.solidFill=color
    ws.add_chart(c2,f"A{chart_anchor_row}"); widths(ws,cw); ws.freeze_panes="B4"

def _sheet_15min(wb, sol):
    n=sol['num']; site=sol.get('site_name','Site'); sdf=sol['site_df']
    total=sol.get('site_summary',{}).get('total_kwh')
    ws=wb.create_sheet(sname(f"Sc {n} – {site} 15min",wb)); ws.sheet_view.showGridLines=False
    htxt=f"  {site}  |  15-min Load Profile"+(f"  |  Annual: {total:,.0f} kWh" if total else "")
    hdr(ws,1,1,3,htxt,size=12); ws.row_dimensions[1].height=26; ws.row_dimensions[2].height=6
    for ci,h in enumerate(["Timestamp","Load (kWh)","Load (kW)"],1):
        c=ws.cell(row=3,column=ci,value=h); c.fill=fill(); c.border=border()
        c.font=Font(name="Arial",bold=True,color=C_WHITE,size=10); c.alignment=aln("center")
    ws.row_dimensions[3].height=18; r=4
    for i,(_,rd) in enumerate(sdf.iterrows()):
        bg=C_LIGHT if (i//96)%2==0 else C_WHITE
        c1=ws.cell(row=r,column=1,value=rd['timestamp']); c1.number_format='DD/MM/YYYY HH:MM'
        c1.fill=fill(bg); c1.font=fnt(size=9); c1.border=border(); c1.alignment=aln("center")
        c2=ws.cell(row=r,column=2,value=rd['load_kwh']); c2.number_format='#,##0.000'
        c2.fill=fill(bg); c2.font=fnt(size=9); c2.border=border(); c2.alignment=aln("right")
        c3=ws.cell(row=r,column=3,value=f"=B{r}*4"); c3.number_format='#,##0.00'
        c3.fill=fill(bg); c3.font=fnt(size=9); c3.border=border(); c3.alignment=aln("right")
        ws.row_dimensions[r].height=12; r+=1
    widths(ws,[18,14,12]); ws.freeze_panes="A4"


# ══════════════════════════════════════════════════════════════════════════════
#  PPT GENERATOR  —  core logic (no tkinter)
# ══════════════════════════════════════════════════════════════════════════════

NS_A   = "http://schemas.openxmlformats.org/drawingml/2006/main"
NS_P   = "http://schemas.openxmlformats.org/presentationml/2006/main"
NS_C   = "http://schemas.openxmlformats.org/drawingml/2006/chart"
NS_R   = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
NS_PKG = "http://schemas.openxmlformats.org/package/2006/relationships"

def _a(tag):  return f"{{{NS_A}}}{tag}"
def _p(tag):  return f"{{{NS_P}}}{tag}"
def _c(tag):  return f"{{{NS_C}}}{tag}"
def _r(tag):  return f"{{{NS_R}}}{tag}"
def _pk(tag): return f"{{{NS_PKG}}}{tag}"

SCENARIO_SLIDE_PAIRS = [
    ("slide6.xml",  "slide7.xml"),
    ("slide8.xml",  "slide9.xml"),
    ("slide10.xml", "slide11.xml"),
]
SCENARIO_CHART_GROUPS = [
    ["chart3.xml",  "chart4.xml",  "chart5.xml",  "chart6.xml"],
    ["chart7.xml",  "chart8.xml",  "chart9.xml",  "chart10.xml"],
    ["chart11.xml", "chart12.xml", "chart13.xml", "chart14.xml"],
]

def _v(val, fallback=""): return val if val is not None else fallback
def _fmt(val, decimals=0, unit=""):
    if val is None or str(val).strip() == "": return "[fill in]"
    try:
        num=float(val); s=f"{num:,.{decimals}f}"
        return f"{s} {unit}".strip() if unit else s
    except: return str(val)
def _fmt_mwh(kwh):
    if not kwh: return "0.0"
    return f"{float(kwh)/1000:,.1f}"
def _fmt_pct(num, den):
    try:
        if not den: return "0.0%"
        return f"{float(num)/float(den)*100:.1f}%"
    except: return "0.0%"

def read_excel_for_ppt(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    cfg = wb["⚙️ Config"]
    def cell(r, c): return cfg.cell(row=r, column=c).value
    info = {
        "project":        str(_v(cell(5,2),"")),
        "location":       str(_v(cell(6,2),"")),
        "date":           str(_v(cell(7,2), datetime.today().strftime("%d %b %Y"))),
        "prepared_by":    str(_v(cell(8,2),"")),
        "client":         str(_v(cell(9,2),"")),
        "client_contact": str(_v(cell(10,2),"")),
    }
    SPEC_KEYS = ["label","pv_kwp","pv_kva","bess_kw","bess_kwh","gen_kva","grid_kva","panels","pv_capex","bess_capex","ppa","epc","pv_life","bess_life"]
    scenarios = []; r = 12
    while r < 300:
        cell_val = cfg.cell(row=r, column=1).value
        if isinstance(cell_val, str) and "Scenario" in cell_val:
            r += 1; spec = {}
            for i, key in enumerate(SPEC_KEYS): spec[key] = cfg.cell(row=r+i, column=2).value
            scenarios.append(spec); r += len(SPEC_KEYS) + 1
        else:
            r += 1
            if r > 200: break
    load_tabs = {}; prod_tabs = {}; load_n = 1; prod_n = 1
    for ws in wb.worksheets:
        t = ws.title
        if "Load" in t and t.startswith("Sc"): load_tabs[load_n] = ws; load_n += 1
        elif "Prod" in t and t.startswith("Sc"): prod_tabs[prod_n] = ws; prod_n += 1
    for i, sc in enumerate(scenarios):
        sn = i + 1
        d = {k: [] for k in ["months","load","grid","genset","battery","solar","stored","curtailed","production"]}
        lws = load_tabs.get(sn)
        if lws:
            row = 4
            while True:
                mv = lws.cell(row=row, column=1).value
                if mv is None or (isinstance(mv, str) and mv.upper().startswith("TOTAL")): break
                d["months"].append(str(mv))
                d["load"].append(float(lws.cell(row=row,column=2).value or 0))
                d["grid"].append(float(lws.cell(row=row,column=3).value or 0))
                d["genset"].append(float(lws.cell(row=row,column=4).value or 0))
                d["battery"].append(float(lws.cell(row=row,column=5).value or 0))
                d["solar"].append(float(lws.cell(row=row,column=6).value or 0))
                row += 1
        pws = prod_tabs.get(sn)
        if pws:
            row = 4
            while True:
                mv = pws.cell(row=row, column=1).value
                if mv is None or (isinstance(mv, str) and mv.upper().startswith("TOTAL")): break
                d["production"].append(float(pws.cell(row=row,column=2).value or 0))
                d["stored"].append(float(pws.cell(row=row,column=4).value or 0))
                d["curtailed"].append(float(pws.cell(row=row,column=5).value or 0))
                row += 1
        for key in ["annual_load","annual_grid","annual_genset","annual_battery","annual_solar","annual_stored","annual_curtailed","annual_production"]:
            src = key.replace("annual_","")
            d[key] = sum(d.get(src, []))
        sc.update(d)
    return info, scenarios

class PptxFiles:
    def __init__(self, template_path):
        self.files = {}
        with zipfile.ZipFile(template_path, 'r') as zf:
            for name in zf.namelist(): self.files[name] = zf.read(name)
    def has(self, path): return path in self.files
    def get_bytes(self, path): return self.files[path]
    def set_bytes(self, path, data): self.files[path] = data
    def delete(self, path): self.files.pop(path, None)
    def get_xml(self, path): return etree.fromstring(self.files[path])
    def set_xml(self, path, root):
        self.files[path] = etree.tostring(root, xml_declaration=True, encoding="UTF-8", standalone=True)
    def save(self, output_path):
        with zipfile.ZipFile(output_path, 'w', zipfile.ZIP_DEFLATED) as zf:
            for name, data in self.files.items(): zf.writestr(name, data)
    def save_to_bytes(self):
        # Write zip verbatim — identical to the original save() except output goes
        # to BytesIO. No XML manipulation: the original working script never touches
        # [Content_Types].xml or chart rels on save, and PowerPoint tolerates orphaned
        # Override entries for removed slides.
        buf = BytesIO()
        with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
            for name, data in self.files.items(): zf.writestr(name, data)
        return buf.getvalue()
    def copy_file(self, src, dst):
        if src in self.files: self.files[dst] = self.files[src]

def replace_text_in_xml(root, replacements):
    all_txbodies = list(root.iter(_p("txBody"))) + list(root.iter(_a("txBody")))
    for txBody in all_txbodies:
        for para in txBody.findall(_a("p")):
            runs = para.findall(_a("r"))
            if not runs: continue
            parts = []
            for run in runs:
                t = run.find(_a("t")); parts.append(t.text or "" if t is not None else "")
            full = "".join(parts); new_full = full
            for old, new in replacements.items(): new_full = new_full.replace(old, str(new))
            if new_full != full:
                t0 = runs[0].find(_a("t"))
                if t0 is not None: t0.text = new_full
                for run in runs[1:]:
                    t = run.find(_a("t"))
                    if t is not None: t.text = ""

def replace_text_in_slide(pf, slide_name, replacements):
    root = pf.get_xml(f"ppt/slides/{slide_name}")
    replace_text_in_xml(root, replacements)
    pf.set_xml(f"ppt/slides/{slide_name}", root)

def rebuild_specs_block(root, sc, scenario_num):
    new_lines = [
        ("Technical installation", True),
        (f"PV capacity (DC)\t{_fmt(sc.get('pv_kwp'))} kWp", False),
        (f"PV inverter capacity\t{_fmt(sc.get('pv_kva'))} kVa", False),
        (f"BESS capacity\t{_fmt(sc.get('bess_kw'))} kW / {_fmt(sc.get('bess_kwh'))} kWh", False),
        (f"Generator\t{_fmt(sc.get('gen_kva'))} kVA", False),
        (f"Grid connection\t{_fmt(sc.get('grid_kva'))} kVA", False),
        ("", False),
        ("Consumption", True),
        (f"Total use\t{_fmt_mwh(sc.get('annual_load',0))} MWh/yr", False),
        (f"From grid\t{_fmt_mwh(sc.get('annual_grid',0))} MWh/yr  ({_fmt_pct(sc.get('annual_grid',0),sc.get('annual_load',1))})", False),
        (f"From solar (direct)\t{_fmt_mwh(sc.get('annual_solar',0))} MWh/yr  ({_fmt_pct(sc.get('annual_solar',0),sc.get('annual_load',1))})", False),
        (f"From BESS\t{_fmt_mwh(sc.get('annual_battery',0))} MWh/yr  ({_fmt_pct(sc.get('annual_battery',0),sc.get('annual_load',1))})", False),
        (f"From generator\t{_fmt_mwh(sc.get('annual_genset',0))} MWh/yr  ({_fmt_pct(sc.get('annual_genset',0),sc.get('annual_load',1))})", False),
        ("", False),
        ("Production", True),
        (f"Total production\t{_fmt_mwh(sc.get('annual_production',0))} MWh/yr", False),
        (f"Direct use\t{_fmt_mwh(sc.get('annual_solar',0))} MWh/yr  ({_fmt_pct(sc.get('annual_solar',0),sc.get('annual_production',1))})", False),
        (f"Stored in BESS\t{_fmt_mwh(sc.get('annual_stored',0))} MWh/yr  ({_fmt_pct(sc.get('annual_stored',0),sc.get('annual_production',1))})", False),
        (f"Curtailed\t{_fmt_mwh(sc.get('annual_curtailed',0))} MWh/yr  ({_fmt_pct(sc.get('annual_curtailed',0),sc.get('annual_production',1))})", False),
    ]
    TAB_STOP_POS = "2874963"
    for sp in root.iter():
        if not sp.tag.endswith("}sp"): continue
        txBody = sp.find(_p("txBody")) or sp.find(_a("txBody"))
        if txBody is None: continue
        all_text = "".join((t.text or "") for t in txBody.iter(_a("t")))
        if "Technical installation" not in all_text and "Installed capacity" not in all_text: continue
        existing_paras = txBody.findall(_a("p"))
        if not existing_paras: continue
        header_tmpl = existing_paras[0]
        body_tmpl   = existing_paras[1] if len(existing_paras) >= 2 else existing_paras[0]
        def ensure_tab_stop(para):
            pPr = para.find(_a("pPr"))
            if pPr is None: pPr = etree.Element(_a("pPr")); para.insert(0, pPr)
            tabLst = pPr.find(_a("tabLst"))
            if tabLst is None: tabLst = etree.SubElement(pPr, _a("tabLst"))
            for t in list(tabLst.findall(_a("tab"))): tabLst.remove(t)
            tab = etree.SubElement(tabLst, _a("tab")); tab.set("pos", TAB_STOP_POS); tab.set("algn", "l")
        def make_para(text, is_header):
            tmpl = header_tmpl if is_header else body_tmpl
            new_para = copy.deepcopy(tmpl)
            runs = new_para.findall(_a("r"))
            if runs:
                t0 = runs[0].find(_a("t"))
                if t0 is not None:
                    t0.text = text
                    if "\t" in text or text.startswith(" ") or text.endswith(" "):
                        t0.set("{http://www.w3.org/XML/1998/namespace}space", "preserve")
                for run in runs[1:]: new_para.remove(run)
            else:
                r_el = etree.SubElement(new_para, _a("r"))
                t_el = etree.SubElement(r_el, _a("t")); t_el.text = text
            if not is_header: ensure_tab_stop(new_para)
            return new_para
        for para in list(txBody.findall(_a("p"))): txBody.remove(para)
        for text, is_header in new_lines: txBody.append(make_para(text, is_header))
        break

def fix_chart_legend(pf, chart_name):
    root = pf.get_xml(f"ppt/charts/{chart_name}")
    legend = root.find(f".//{_c('legend')}")
    if legend is not None:
        layout = legend.find(_c("layout"))
        if layout is not None: legend.remove(layout)
    pf.set_xml(f"ppt/charts/{chart_name}", root)

def update_chart_title(pf, chart_name, new_title):
    root = pf.get_xml(f"ppt/charts/{chart_name}")
    title_el = root.find(f".//{_c('title')}")
    if title_el is None: pf.set_xml(f"ppt/charts/{chart_name}", root); return
    rich = title_el.find(f".//{_c('rich')}")
    if rich is None: pf.set_xml(f"ppt/charts/{chart_name}", root); return
    para = rich.find(_a("p"))
    if para is None: pf.set_xml(f"ppt/charts/{chart_name}", root); return
    existing_runs = para.findall(_a("r"))
    rpr = None
    if existing_runs:
        rpr = existing_runs[0].find(_a("rPr"))
        for r_el in existing_runs: para.remove(r_el)
    end_rpr = para.find(_a("endParaRPr"))
    if end_rpr is not None: para.remove(end_rpr)
    new_r = etree.SubElement(para, _a("r"))
    if rpr is not None: new_r.append(copy.deepcopy(rpr))
    new_t = etree.SubElement(new_r, _a("t")); new_t.text = new_title
    pf.set_xml(f"ppt/charts/{chart_name}", root)

def update_chart_data(pf, chart_name, series_values_list, categories=None):
    root = pf.get_xml(f"ppt/charts/{chart_name}")
    sers = root.findall(f".//{_c('ser')}")
    for ser_idx, ser in enumerate(sers):
        if ser_idx >= len(series_values_list): break
        new_vals = series_values_list[ser_idx]
        val_el = ser.find(_c("val"))
        if val_el is not None:
            numRef = val_el.find(_c("numRef"))
            if numRef is not None:
                numCache = numRef.find(_c("numCache"))
                if numCache is not None:
                    ptCount = numCache.find(_c("ptCount"))
                    if ptCount is not None: ptCount.set("val", str(len(new_vals)))
                    for pt in numCache.findall(_c("pt")): numCache.remove(pt)
                    for i, val in enumerate(new_vals):
                        pt = etree.SubElement(numCache, _c("pt")); pt.set("idx", str(i))
                        v = etree.SubElement(pt, _c("v")); v.text = str(round(float(val), 4))
        if categories is not None:
            cat_el = ser.find(_c("cat"))
            if cat_el is not None:
                strRef = cat_el.find(_c("strRef"))
                if strRef is not None:
                    strCache = strRef.find(_c("strCache"))
                    if strCache is not None:
                        ptCount = strCache.find(_c("ptCount"))
                        if ptCount is not None: ptCount.set("val", str(len(categories)))
                        for pt in strCache.findall(_c("pt")): strCache.remove(pt)
                        for i, cat in enumerate(categories):
                            pt = etree.SubElement(strCache, _c("pt")); pt.set("idx", str(i))
                            v = etree.SubElement(pt, _c("v")); v.text = str(cat)
    pf.set_xml(f"ppt/charts/{chart_name}", root)

def _build_dLbl_right(color_hex, val, is_scheme=False):
    dLbls = etree.Element(_c("dLbls"))
    if val == 0.0:
        for tag in ("showLegendKey","showVal","showCatName","showSerName","showPercent","showBubbleSize"):
            etree.SubElement(dLbls, _c(tag)).set("val", "0")
        return dLbls
    dLbl = etree.SubElement(dLbls, _c("dLbl"))
    etree.SubElement(dLbl, _c("idx")).set("val", "0")
    # No manualLayout — use outEnd so PowerPoint auto-positions the label
    # cleanly at the outside end of each bar segment. No clipping, works for
    # any bar size or data value.
    spPr = etree.SubElement(dLbl, _c("spPr"))
    etree.SubElement(spPr, _a("noFill"))
    ln = etree.SubElement(spPr, _a("ln")); etree.SubElement(ln, _a("noFill"))
    etree.SubElement(spPr, _a("effectLst"))
    txPr = etree.SubElement(dLbl, _c("txPr"))
    bodyPr = etree.SubElement(txPr, _a("bodyPr"))
    bodyPr.set("wrap","none"); bodyPr.set("lIns","0"); bodyPr.set("tIns","0")
    bodyPr.set("rIns","0"); bodyPr.set("bIns","0"); bodyPr.set("anchor","ctr")
    etree.SubElement(bodyPr, _a("spAutoFit")); etree.SubElement(txPr, _a("lstStyle"))
    p = etree.SubElement(txPr, _a("p"))
    pPr = etree.SubElement(p, _a("pPr")); defRPr = etree.SubElement(pPr, _a("defRPr"))
    defRPr.set("sz","1200"); defRPr.set("b","0")
    solidFill = etree.SubElement(defRPr, _a("solidFill"))
    if is_scheme:
        schemeClr = etree.SubElement(solidFill, _a("schemeClr")); schemeClr.set("val", color_hex)
        etree.SubElement(schemeClr, _a("lumMod")).set("val","50000"); etree.SubElement(schemeClr, _a("lumOff")).set("val","50000")
    else:
        etree.SubElement(solidFill, _a("srgbClr")).set("val", color_hex)
    etree.SubElement(defRPr, _a("latin")).set("typeface","Poppins")
    etree.SubElement(defRPr, _a("cs")).set("typeface","Poppins")
    etree.SubElement(p, _a("endParaRPr")).set("lang","en-US")
    etree.SubElement(dLbl, _c("dLblPos")).set("val","outEnd")
    for tag, v in [("showLegendKey","0"),("showVal","1"),("showCatName","0"),("showSerName","0"),("showPercent","0"),("showBubbleSize","0"),("showLeaderLines","0")]:
        etree.SubElement(dLbl, _c(tag)).set("val", v)
    for tag, v in [("showLegendKey","0"),("showVal","1"),("showCatName","0"),("showSerName","0"),("showPercent","0"),("showBubbleSize","0"),("showLeaderLines","0")]:
        etree.SubElement(dLbls, _c(tag)).set("val", v)
    return dLbls

def clean_annual_chart_labels(pf, chart_name):
    root = pf.get_xml(f"ppt/charts/{chart_name}")
    chart_dLbls = root.find(f".//{_c('plotArea')}//{_c('dLbls')}")
    if chart_dLbls is not None:
        sv = chart_dLbls.find(_c("showVal"))
        if sv is not None: sv.set("val","0")
    sers = root.findall(f".//{_c('ser')}")
    for ser in sers:
        num_vs = ser.findall(f".//{_c('numRef')}//{_c('v')}")
        try: val = float(num_vs[0].text or 0) if num_vs else 0.0
        except ValueError: val = 0.0
        srgb = ser.find(f".//{_c('spPr')}//{_a('srgbClr')}")
        scheme = ser.find(f".//{_c('spPr')}//{_a('schemeClr')}")
        if srgb is not None: color_hex = srgb.get("val","404040"); is_scheme = False
        elif scheme is not None: color_hex = scheme.get("val","tx1"); is_scheme = True
        else: color_hex = "404040"; is_scheme = False
        old_dLbls = ser.find(_c("dLbls"))
        if old_dLbls is not None: ser.remove(old_dLbls)
        ser.append(_build_dLbl_right(color_hex, val, is_scheme))
    pf.set_xml(f"ppt/charts/{chart_name}", root)

def hide_zero_series_in_legend(pf, chart_name):
    root = pf.get_xml(f"ppt/charts/{chart_name}")
    legend = root.find(f".//{_c('legend')}")
    if legend is None: pf.set_xml(f"ppt/charts/{chart_name}", root); return
    for le in legend.findall(_c("legendEntry")): legend.remove(le)
    sers = root.findall(f".//{_c('ser')}")
    for ser in sers:
        idx_el = ser.find(_c("idx")); idx = idx_el.get("val","0") if idx_el is not None else "0"
        num_vs = ser.findall(f".//{_c('numRef')}//{_c('v')}")
        vals = []
        for v in num_vs:
            try: vals.append(float(v.text or 0))
            except ValueError: vals.append(0.0)
        if not vals or all(v == 0.0 for v in vals):
            le = etree.Element(_c("legendEntry"))
            etree.SubElement(le, _c("idx")).set("val", idx)
            etree.SubElement(le, _c("delete")).set("val","1")
            legend.insert(0, le)
    pf.set_xml(f"ppt/charts/{chart_name}", root)

def update_scenario_charts(pf, chart_group, sc):
    months = sc.get("months") or MONTHS; label = sc.get("label") or "Annual"
    for chart_name in [chart_group[0], chart_group[2]]: fix_chart_legend(pf, chart_name)
    update_chart_title(pf, chart_group[0], "Energy use")
    update_chart_title(pf, chart_group[2], "Solar PV production")
    update_chart_data(pf, chart_group[0], [sc.get("genset",[0]*12), sc.get("grid",[0]*12), sc.get("battery",[0]*12), sc.get("solar",[0]*12)], categories=months)
    update_chart_data(pf, chart_group[1], [[sc.get("annual_genset",0)],[sc.get("annual_grid",0)],[sc.get("annual_battery",0)],[sc.get("annual_solar",0)]], categories=[label])
    clean_annual_chart_labels(pf, chart_group[1])
    hide_zero_series_in_legend(pf, chart_group[0])
    update_chart_data(pf, chart_group[2], [sc.get("solar",[0]*12), sc.get("stored",[0]*12), sc.get("curtailed",[0]*12)], categories=months)
    update_chart_data(pf, chart_group[3], [[sc.get("annual_solar",0)],[sc.get("annual_stored",0)],[sc.get("annual_curtailed",0)]], categories=[label])
    clean_annual_chart_labels(pf, chart_group[3])
    hide_zero_series_in_legend(pf, chart_group[2])

def remove_slide_from_presentation(pf, slide_filename):
    prs_root = pf.get_xml("ppt/presentation.xml")
    prs_rels_root = pf.get_xml("ppt/_rels/presentation.xml.rels")
    rid_to_remove = None
    for rel in prs_rels_root.findall(_pk("Relationship")):
        if rel.get("Target","").endswith(slide_filename):
            rid_to_remove = rel.get("Id"); prs_rels_root.remove(rel); break
    sldIdLst = prs_root.find(_p("sldIdLst"))
    for sldId in sldIdLst.findall(_p("sldId")):
        if sldId.get(_r("id")) == rid_to_remove: sldIdLst.remove(sldId); break
    slide_rels_path = f"ppt/slides/_rels/{slide_filename}.rels"
    if pf.has(slide_rels_path):
        slide_rels_root = pf.get_xml(slide_rels_path)
        for rel in slide_rels_root.findall(_pk("Relationship")):
            rel_type = rel.get("Type",""); target = rel.get("Target","")
            if "chart" in rel_type:
                chart_fname = os.path.basename(target)
                pf.delete(f"ppt/charts/{chart_fname}"); pf.delete(f"ppt/charts/_rels/{chart_fname}.rels")
            elif "notesSlide" in rel_type:
                notes_fname = os.path.basename(target)
                pf.delete(f"ppt/notesSlides/{notes_fname}"); pf.delete(f"ppt/notesSlides/_rels/{notes_fname}.rels")
        pf.delete(slide_rels_path)
    pf.delete(f"ppt/slides/{slide_filename}")
    pf.set_xml("ppt/presentation.xml", prs_root)
    pf.set_xml("ppt/_rels/presentation.xml.rels", prs_rels_root)

def duplicate_scenario_pair(pf, source_pair_idx, new_scenario_num, insert_before_filename):
    src_tech, src_obs = SCENARIO_SLIDE_PAIRS[source_pair_idx]
    src_charts = SCENARIO_CHART_GROUPS[source_pair_idx]
    existing_slides = [n for n in pf.files if re.match(r"ppt/slides/slide\d+\.xml$", n)]
    max_slide = max(int(re.search(r"slide(\d+)", n).group(1)) for n in existing_slides)
    existing_charts = [n for n in pf.files if re.match(r"ppt/charts/chart\d+\.xml$", n)]
    max_chart = max(int(re.search(r"chart(\d+)", n).group(1)) for n in existing_charts) if existing_charts else 0
    new_tech_num = max_slide + 1; new_obs_num = max_slide + 2
    new_tech_name = f"slide{new_tech_num}.xml"; new_obs_name = f"slide{new_obs_num}.xml"
    new_chart_nums = [max_chart+1, max_chart+2, max_chart+3, max_chart+4]
    new_chart_names = [f"chart{n}.xml" for n in new_chart_nums]
    for src_c, dst_c in zip(src_charts, new_chart_names):
        pf.copy_file(f"ppt/charts/{src_c}", f"ppt/charts/{dst_c}")
        src_rels = f"ppt/charts/_rels/{src_c}.rels"
        if pf.has(src_rels): pf.copy_file(src_rels, f"ppt/charts/_rels/{dst_c}.rels")
    def build_tech_rels(new_chart_names):
        src_rels_path = f"ppt/slides/_rels/{src_tech}.rels"
        if not pf.has(src_rels_path): return None
        rels_root = pf.get_xml(src_rels_path); chart_rel_ids = ["rId3","rId4","rId5","rId6"]; chart_idx = 0
        for rel in rels_root.findall(_pk("Relationship")):
            if "chart" in rel.get("Type",""):
                if chart_idx < len(new_chart_names): rel.set("Target", f"../charts/{new_chart_names[chart_idx]}"); chart_idx += 1
        return rels_root
    tech_root = pf.get_xml(f"ppt/slides/{src_tech}"); pf.set_xml(f"ppt/slides/{new_tech_name}", tech_root)
    tech_rels = build_tech_rels(new_chart_names)
    if tech_rels is not None: pf.set_xml(f"ppt/slides/_rels/{new_tech_name}.rels", tech_rels)
    obs_root = pf.get_xml(f"ppt/slides/{src_obs}"); pf.set_xml(f"ppt/slides/{new_obs_name}", obs_root)
    src_obs_rels = f"ppt/slides/_rels/{src_obs}.rels"
    if pf.has(src_obs_rels): pf.copy_file(src_obs_rels, f"ppt/slides/_rels/{new_obs_name}.rels")
    prs_root = pf.get_xml("ppt/presentation.xml")
    prs_rels_root = pf.get_xml("ppt/_rels/presentation.xml.rels")
    existing_rids = [rel.get("Id","") for rel in prs_rels_root.findall(_pk("Relationship"))]
    max_rid = max((int(re.search(r"rId(\d+)", rid).group(1)) for rid in existing_rids if re.search(r"rId(\d+)", rid)), default=0)
    new_rid_tech = f"rId{max_rid+1}"; new_rid_obs = f"rId{max_rid+2}"
    REL_SLIDE_TYPE = "http://schemas.openxmlformats.org/officeDocument/2006/relationships/slide"
    for rid, fname in [(new_rid_tech, new_tech_name), (new_rid_obs, new_obs_name)]:
        rel = etree.SubElement(prs_rels_root, _pk("Relationship"))
        rel.set("Id", rid); rel.set("Type", REL_SLIDE_TYPE); rel.set("Target", f"slides/{fname}")
    sldIdLst = prs_root.find(_p("sldIdLst"))
    existing_ids = [sldId.get("id") for sldId in sldIdLst.findall(_p("sldId"))]
    max_id = max(int(x) for x in existing_ids if x and x.isdigit())
    insert_pos = len(sldIdLst)
    for i, sldId in enumerate(sldIdLst.findall(_p("sldId"))):
        rid = sldId.get(_r("id"))
        for rel in prs_rels_root.findall(_pk("Relationship")):
            if rel.get("Id") == rid and rel.get("Target","").endswith(insert_before_filename):
                insert_pos = i; break
    def make_sld_id(num_id, rid):
        el = etree.Element(_p("sldId")); el.set("id", str(num_id)); el.set(_r("id"), rid); return el
    sldIdLst.insert(insert_pos, make_sld_id(max_id+2, new_rid_obs))
    sldIdLst.insert(insert_pos, make_sld_id(max_id+1, new_rid_tech))
    pf.set_xml("ppt/presentation.xml", prs_root)
    pf.set_xml("ppt/_rels/presentation.xml.rels", prs_rels_root)
    return new_tech_name, new_obs_name, new_chart_names

def update_cover(pf, info):
    client = info["client"] or "[Client]"; location = info["location"] or "[Location]"; date = info["date"]
    replace_text_in_slide(pf, "slide1.xml", {"Date March 16th 2026":date,"March 16th 2026":date,"6AM NV":client,"[client]":client,"[Client]":client,"[location]":location,"[Location]":location})

def update_scenario_technical(pf, slide_name, chart_names, sc, scenario_num):
    label = sc.get("label") or ""; title = f"Scenario {scenario_num}: {label}" if label else f"Scenario {scenario_num}"
    root = pf.get_xml(f"ppt/slides/{slide_name}")
    replace_text_in_xml(root, {
        "Scenario 1: Solution 1 with [X] profile":title,"Scenario 2: Solution 1 with [X] profile":title,
        "Scenario 3: Solution 2 with [Y] profile":title,"Solution 1":label or f"Scenario {scenario_num}",
        "Solution 2":label or f"Scenario {scenario_num}","Solution 3":label or f"Scenario {scenario_num}",
        "[X] profile":label or f"Scenario {scenario_num}","[Y] profile":label or f"Scenario {scenario_num}",
        "Note: This solution is insufficient. It leads to an overload of the grid connection for 129 days per year, with a max. overload of 110 kW and a total overload of 57.3 MWh.":"[Add any technical notes or caveats here]",
    })
    pf.set_xml(f"ppt/slides/{slide_name}", root)
    root = pf.get_xml(f"ppt/slides/{slide_name}")
    rebuild_specs_block(root, sc, scenario_num)
    pf.set_xml(f"ppt/slides/{slide_name}", root)
    update_scenario_charts(pf, chart_names, sc)

def update_scenario_observations(pf, slide_name, sc, scenario_num):
    label = sc.get("label") or f"Scenario {scenario_num}"
    replace_text_in_slide(pf, slide_name, {
        "Scenario 1: Solution 1 observations":f"Scenario {scenario_num}: {label} – observations",
        "Scenario 2: Solution 2 observations":f"Scenario {scenario_num}: {label} – observations",
        "Scenario 3: Solution 2 observations":f"Scenario {scenario_num}: {label} – observations",
        "Solution 1 observations":f"{label} – observations","Solution 2 observations":f"{label} – observations",
    })

def generate_ppt(excel_path, template_path, progress=None):
    def log(msg):
        if progress: progress(msg)
    log("Reading Excel data...")
    info, scenarios = read_excel_for_ppt(excel_path)
    n = len(scenarios)
    log(f"Found {n} scenario(s)...")
    log("Loading template...")
    pf = PptxFiles(template_path)
    template_pairs = len(SCENARIO_SLIDE_PAIRS)
    active_pairs = list(SCENARIO_SLIDE_PAIRS); active_charts = list(SCENARIO_CHART_GROUPS)
    # Normalise chart formatting: overwrite any earlier template pair's chart bytes
    # with the last pair's bytes before updating data. This makes all output slides
    # share the same chart structure/styles as the last (always-correct) template slide.
    # Uses existing chart file names so [Content_Types].xml never needs touching.
    last_charts = SCENARIO_CHART_GROUPS[template_pairs - 1]
    for i in range(min(n, template_pairs - 1)):
        for src_c, dst_c in zip(last_charts, SCENARIO_CHART_GROUPS[i]):
            pf.copy_file(f"ppt/charts/{src_c}", f"ppt/charts/{dst_c}")
            src_rels = f"ppt/charts/_rels/{src_c}.rels"
            dst_rels = f"ppt/charts/_rels/{dst_c}.rels"
            if pf.has(src_rels): pf.copy_file(src_rels, dst_rels)
    if n > template_pairs:
        for extra_i in range(template_pairs, n):
            new_tech, new_obs, new_charts = duplicate_scenario_pair(pf, template_pairs-1, extra_i+1, "ZZZNOMATCH.xml")
            active_pairs.append((new_tech, new_obs)); active_charts.append(new_charts)
    for i in range(n):
        sc = scenarios[i]; tech_slide, obs_slide = active_pairs[i]; chart_group = active_charts[i]
        log(f"Updating Scenario {i+1}: {sc.get('label') or '(unnamed)'}...")
        update_scenario_technical(pf, tech_slide, chart_group, sc, i+1)
    if n < template_pairs:
        for i in range(template_pairs-1, n-1, -1):
            tech_slide, obs_slide = SCENARIO_SLIDE_PAIRS[i]
            remove_slide_from_presentation(pf, obs_slide)
            remove_slide_from_presentation(pf, tech_slide)
    update_cover(pf, info)
    # Build keep set: all non-scenario slides (cover, appendix, etc.) + active tech slides only
    all_scenario_slides = set(s for pair in SCENARIO_SLIDE_PAIRS for s in pair)
    keep_slides = set(s for pair in SCENARIO_SLIDE_PAIRS for s in pair if s not in all_scenario_slides)
    # Keep non-scenario slides (template slides outside the scenario pairs)
    keep_slides = set()
    for tech_slide, obs_slide in SCENARIO_SLIDE_PAIRS:
        pass  # scenario slides handled below
    prs_rels_root = pf.get_xml("ppt/_rels/presentation.xml.rels")
    rels_map = {r.get("Id"): r.get("Target","").split("/")[-1]
                for r in prs_rels_root.findall(f"{{{NS_PKG}}}Relationship")
                if "slide" in r.get("Type","") and "slideMaster" not in r.get("Type","") and "slideLayout" not in r.get("Type","")}
    # Keep: non-scenario slides + active tech slides (obs slides are always removed)
    for fname in rels_map.values():
        if fname not in all_scenario_slides:
            keep_slides.add(fname)  # cover, appendix, closing slides — always keep
    for tech_slide, _obs_slide in active_pairs[:n]:
        keep_slides.add(tech_slide)  # active scenario tech slides
    slides_to_remove = [fname for fname in rels_map.values() if fname not in keep_slides]
    for fname in slides_to_remove: remove_slide_from_presentation(pf, fname)
    log("Building presentation...")
    return pf.save_to_bytes()


# ══════════════════════════════════════════════════════════════════════════════
#  STREAMLIT UI
# ══════════════════════════════════════════════════════════════════════════════

import base64 as _b64
_logo_path = os.path.join(SCRIPT_DIR, "bluepeak_logo.png")
_logo_b64 = ""
if os.path.exists(_logo_path):
    with open(_logo_path, "rb") as _f: _logo_b64 = _b64.b64encode(_f.read()).decode()

st.set_page_config(page_title="Bluepeak | Wattix Tools", page_icon="⚡", layout="centered")

st.markdown(f"""
<style>
/* ── Page background ── */
.stApp {{ background-color: #f0f6fb; }}

/* ── Header banner ── */
.bp-header {{
    background: linear-gradient(135deg, #012540 0%, #034C81 100%);
    padding: 18px 28px 16px 28px; border-radius: 10px; margin-bottom: 24px;
    box-shadow: 0 3px 10px rgba(1,37,64,0.25);
    display: flex; align-items: center; gap: 20px;
}}
.bp-header img {{ height: 38px; filter: brightness(0) invert(1); }}
.bp-header-text {{ color: white; font-size: 13px; opacity: 0.75; margin-top: 3px; font-family: Arial; }}

/* ── Tabs ── */
.stTabs [data-baseweb="tab-list"] {{ gap: 4px; border-bottom: 2px solid #B2E0FF; }}
.stTabs [data-baseweb="tab"] {{
    background-color: #e6f2f8; border-radius: 6px 6px 0 0;
    color: #012540; font-weight: 600; padding: 8px 20px;
    border: 1px solid #B2E0FF; border-bottom: none;
}}
.stTabs [aria-selected="true"] {{
    background-color: #034C81 !important; color: white !important;
    border-color: #034C81 !important;
}}
.stTabs [data-baseweb="tab"]:hover {{ background-color: #1DB4C5; color: white; }}

/* ── Primary buttons ── */
div.stButton > button[kind="primary"] {{
    background-color: #034C81; color: white; border: none;
    border-radius: 6px; font-weight: 600; letter-spacing: 0.3px;
    transition: background 0.2s;
}}
div.stButton > button[kind="primary"]:hover {{ background-color: #1DB4C5; color: white; }}
div.stButton > button[kind="primary"]:active {{ background-color: #012540; }}

/* ── Secondary / clear button ── */
div.stButton > button[kind="secondary"] {{
    border: 1px solid #034C81; color: #034C81; border-radius: 6px; font-weight: 500;
}}
div.stButton > button[kind="secondary"]:hover {{ background-color: #e6f2f8; }}

/* ── Download button ── */
div.stDownloadButton > button {{
    background-color: #15A781; color: white; border: none;
    border-radius: 6px; font-weight: 600;
}}
div.stDownloadButton > button:hover {{ background-color: #0d8a6a; color: white; }}

/* ── Inputs ── */
div[data-testid="stTextInput"] input:focus {{ border-color: #034C81; box-shadow: 0 0 0 1px #034C81; }}

/* ── Divider ── */
hr {{ border-top: 1px solid #B2E0FF !important; }}

/* ── Section labels ── */
.bp-section {{ color: #012540; font-weight: 700; font-size: 14px;
    border-left: 3px solid #1DB4C5; padding-left: 8px; margin: 4px 0 10px 0; }}
</style>

<div class="bp-header">
  {"<img src='data:image/png;base64," + _logo_b64 + "' />" if _logo_b64 else "<span style='color:white;font-size:20px;font-weight:700;'>⚡ bluepeak.energy</span>"}
  <div><div class="bp-header-text">Wattix Tools</div></div>
</div>
""", unsafe_allow_html=True)

tab1, tab2 = st.tabs(["📊 Excel Generator", "📑 PowerPoint Generator"])


# ── Tab 1: Excel Generator ────────────────────────────────────────────────────
with tab1:
    if "xl_uploader_key" not in st.session_state: st.session_state["xl_uploader_key"] = 0

    st.markdown('<div class="bp-section">1 — Upload your Wattix files</div>', unsafe_allow_html=True)
    st.caption("CSV or XLSX — load, production, generation. Upload all files for your scenarios at once.")

    col_up, col_clr = st.columns([6, 1])
    with col_up:
        uploaded_files = st.file_uploader(
            "Wattix files", accept_multiple_files=True, type=["csv","xlsx"],
            label_visibility="collapsed", key=f"excel_uploader_{st.session_state['xl_uploader_key']}"
        )
    with col_clr:
        st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
        if st.button("🗑️ Clear", key="xl_clear", help="Remove all uploaded files"):
            st.session_state["xl_uploader_key"] += 1
            st.rerun()

    if uploaded_files:
        with tempfile.TemporaryDirectory() as tmpdir:
            paths = []
            for f in uploaded_files:
                fpath = os.path.join(tmpdir, f.name)
                with open(fpath, "wb") as out: out.write(f.getbuffer())
                paths.append(fpath)
            lf, pf_dict, gf, stf, sf = classify_files(paths)
            all_keys = sorted(set(list(lf)+list(pf_dict)+list(gf)), key=str)
        if all_keys:
            st.success(f"✓  {len(all_keys)} scenario(s) detected  |  {len(lf)} load  ·  {len(pf_dict)} production  ·  {len(gf)} generation  ·  {len(sf)} site profile(s)")
        else:
            st.warning("No Wattix scenarios detected. Make sure filenames contain 'wattix' and 'load' or 'production'.")

    st.divider()
    st.markdown('<div class="bp-section">2 — Project details</div>', unsafe_allow_html=True)
    c1, c2, c3 = st.columns(3)
    with c1: project = st.text_input("Project name", placeholder="e.g. VWLD Amstelveen", key="xl_project")
    with c2: location = st.text_input("Location", placeholder="e.g. Amstelveen, NL", key="xl_location")
    with c3: report_date = st.text_input("Report date", value=datetime.today().strftime("%d %b %Y"), key="xl_date")

    st.divider()
    st.markdown('<div class="bp-section">3 — Generate</div>', unsafe_allow_html=True)

    if st.button("⚡  Generate Excel", type="primary", use_container_width=True):
        if not uploaded_files:
            st.error("Please upload your Wattix files first.")
        else:
            with st.spinner("Processing files…"):
                with tempfile.TemporaryDirectory() as tmpdir:
                    paths = []
                    for f in uploaded_files:
                        fpath = os.path.join(tmpdir, f.name)
                        with open(fpath, "wb") as out: out.write(f.getbuffer())
                        paths.append(fpath)
                    lf, pf_dict, gf, stf, sf = classify_files(paths)
                    all_keys = sorted(set(list(lf)+list(pf_dict)+list(gf)), key=str)
                    if not all_keys:
                        st.error("No scenarios detected — check your filenames.")
                    else:
                        site_keys = list(sf.keys())
                        stored_map = _resolve_stored(stf, all_keys)
                        solutions = [load_solution(k, lf.get(k), pf_dict.get(k), gf.get(k), stored_map.get(k),
                                     sf.get(site_keys[0]) if len(site_keys)==1 else None) for k in all_keys]
                        proj_clean = re.sub(r'[^\w\s-]','', project or "Bluepeak").strip().replace(' ','_')
                        outname = f"{proj_clean}_Wattix_Output.xlsx"
                        outpath = os.path.join(tmpdir, outname)
                        build_excel(solutions, project or "Bluepeak Project", location or "", report_date, outpath)
                        with open(outpath, "rb") as f: excel_bytes = f.read()
            st.success(f"✓  Done — {len(all_keys)} scenario(s) processed.")
            st.download_button(
                label="📥  Download Excel",
                data=excel_bytes,
                file_name=outname,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )


# ── Tab 2: PowerPoint Generator ───────────────────────────────────────────────
with tab2:
    st.markdown('<div class="bp-section">1 — Upload your Bluepeak Excel output</div>', unsafe_allow_html=True)
    st.caption("Upload the Excel file generated by the Excel Generator tab.")

    excel_file = st.file_uploader(
        "Bluepeak Excel file", type=["xlsx"],
        label_visibility="collapsed", key="ppt_uploader"
    )

    if excel_file:
        st.success(f"✓  {excel_file.name} loaded")

    st.divider()
    st.markdown('<div class="bp-section">2 — Generate</div>', unsafe_allow_html=True)

    template_ok = os.path.exists(TEMPLATE_PATH)
    if not template_ok:
        st.warning("⚠️  Template not found — make sure `Bluepeak_Template.pptx` is in the same folder as `app.py`.")

    if st.button("📑  Generate PowerPoint", type="primary", use_container_width=True, disabled=not template_ok):
        if not excel_file:
            st.error("Please upload your Bluepeak Excel file first.")
        else:
            log_lines = []
            def progress(msg): log_lines.append(msg)

            with st.spinner("Generating presentation…"):
                with tempfile.TemporaryDirectory() as tmpdir:
                    excel_path = os.path.join(tmpdir, excel_file.name)
                    with open(excel_path, "wb") as f: f.write(excel_file.getbuffer())
                    try:
                        ppt_bytes = generate_ppt(excel_path, TEMPLATE_PATH, progress=progress)
                        proj_name = excel_file.name.replace(".xlsx","").replace("_Wattix_Output","")
                        outname = f"{proj_name}_Presentation.pptx"
                        st.success("✓  Presentation ready.")
                        st.download_button(
                            label="📥  Download PowerPoint",
                            data=ppt_bytes,
                            file_name=outname,
                            mime="application/vnd.openxmlformats-officedocument.presentationml.presentation",
                            use_container_width=True
                        )
                    except Exception as e:
                        import traceback
                        st.error(f"Error: {e}")
                        st.code(traceback.format_exc())
